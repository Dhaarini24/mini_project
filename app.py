from flask import Flask, request, jsonify, render_template,redirect
import pandas as pd
import pickle
import numpy as np
import cv2
from keras.models import load_model
from sklearn.preprocessing import MinMaxScaler
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import train_test_split
import os
from openpyxl import Workbook,load_workbook
app = Flask(__name__)

class SoilFertilityPredictor:
    def __init__(self):
        self.data = None
        self.X = None
        self.Y = None
        self.scaler = MinMaxScaler()
        self.forestRegressor = RandomForestRegressor(criterion='squared_error', max_depth=8, n_estimators=10, random_state=0)

    def load_data(self,processed_data_set):
        self.data = pd.read_csv(r"processed_data_set.csv")
        self.X, self.Y = self.data[self.data.columns[1:]], self.data['Vegetation Cover']
        self.preprocess_data()
        self.train_model()

    def preprocess_data(self):
        X_scaled = self.scaler.fit_transform(self.X.values)
        self.X = X_scaled[:-1]
        self.Y = self.Y[:-1]

    def train_model(self, test_size=0.1, random_state=43):
        X_train, X_test, Y_train, Y_test = train_test_split(self.X, self.Y, test_size=test_size, random_state=random_state)
        self.forestRegressor.fit(X_train, Y_train)
        self.calculate_accuracy(X_test, Y_test)

    def predict_fertility(self, input_data):
        prediction = self.forestRegressor.predict([input_data])
        return prediction[0]

    def evaluate_fertility(self, prediction, nutrient_values):
        a0, a1, a2, a3, a4, a5, a6, a7, a8, a9, a10, a11, a12, a13 = nutrient_values
        text = ""
        if prediction < 90:
            if a0 < 12.75 or a2 < 47 or a8 < 0.6 or a3 < 15 or a6 < 0.28 or a10 < 1:
                text = "Your Soil is less fertile. You may try increasing these nutrients: "
                if a0 < 12.75:
                    text += "NO3 "
                if a2 < 47:
                    text += "P "
                if a8 < 0.6:
                    text += "Zn "
                if a3 < 15:
                    text += "K "
                if a6 < 0.28:
                    text += "Organic Matter "
                if a10 < 1:
                    text += "Fe "
        elif prediction >= 90:
            text = "Your soil is highly fertile."
        return text

    def calculate_accuracy(self, X_test, Y_test):
        accuracy = self.forestRegressor.score(X_test, Y_test)
        print(f"Model Accuracy: {accuracy}")

predictor = SoilFertilityPredictor()
predictor.load_data(r"dataset.txt")
# Load the pre-trained model
with open('new_rf_model.pickle', 'rb') as f:
    model = pickle.load(f)

class_labels = ['apple', 'banana', 'blackgram', 'chickpea', 'coconut', 'coffee',
                'cotton', 'grapes', 'jute', 'kidneybeans', 'lentil', 'maize',
                'mango', 'mothbeans', 'mungbean', 'muskmelon', 'orange', 'papaya',
                'pigeonpeas', 'pomegranate', 'rice', 'watermelon']

model_dis = load_model('my_model.h5')
IMAGE_SIZE = 64

def read_and_resize_image(filepath, image_size):
    img = cv2.imread(filepath)
    resized_img = cv2.resize(img, (image_size, image_size))
    resized_img = resized_img.astype('float32') / 255.0
    return resized_img

def predict_disease(image_path):
    input_image = read_and_resize_image(image_path, IMAGE_SIZE)
    input_image = np.expand_dims(input_image, axis=0)
    predictions = model_dis.predict(input_image)
    predicted_class_index = np.argmax(predictions)
    disease_class = ['Apple_scab','Apple_black_rot','Apple_cedar_apple_rust','Apple_healthy','Background_without_leaves','Blueberry_healthy','Cherry_powdery_mildew','Cherry_healthy','Corn_gray_leaf_spot','Corn_common_rust','Corn_northern_leaf_blight','Corn_healthy','Grape_black_rot','Grape_black_measles','Grape_leaf_blight','Grape_healthy','Orange_haunglongbing','Peach_bacterial_spot','Peach_healthy','Pepper_bacterial_spot','Pepper_healthy','Potato_early_blight','Potato_healthy','Potato_late_blight','Raspberry_healthy','Soybean_healthy','Squash_powdery_mildew','Strawberry_healthy','Strawberry_leaf_scorch','Tomato_bacterial_spot','Tomato_early_blight','Tomato_healthy','Tomato_late_blight','Tomato_leaf_mold','Tomato_septoria_leaf_spot','Tomato_spider_mites_two-spotted_spider_mite','Tomato_target_spot','Tomato_mosaic_virus','Tomato_yellow_leaf_curl_virus']
    predicted_disease = disease_class[predicted_class_index]
    return predicted_disease

@app.route('/')
def index():
    return render_template('landing.html', header='templates\header.html', footer='templates/footer.html')

@app.route('/contact')
def contact():
    return render_template('contact_us.html')

@app.route('/faq')
def faq():
    return render_template('faq.html')
@app.route('/suggestion')
def suggestion():
    return render_template('index.html')

@app.route('/predict', methods=['POST'])
def predict():
    data = request.get_json(force=True)
    input_data = pd.Series({
        'N': data['N'],
        'P': data['P'],
        'K': data['K'],
        'temperature': data['temperature'],
        'humidity': data['humidity'],
        'ph': data['ph'],
        'rainfall': data['rainfall']
    })
    prediction = model.predict([input_data])[0]
    return jsonify({'crop': class_labels[prediction]})

@app.route('/soil_pred')
def soil_pred():
    return render_template('soil_pred.html')

@app.route('/predict_fertility', methods=['POST'])
def predict_fertility():
    data = request.form  # Assuming form data is being sent from index.html
    nutrient_values = [
        float(data['NO3']),
        float(data['NH4']),
        float(data['P']),
        float(data['K']),
        float(data['SO4']),
        float(data['B']),
        float(data['Organic_Matter']),
        float(data['pH']),
        float(data['Zn']),
        float(data['Cu']),
        float(data['Fe']),
        float(data['Ca']),
        float(data['Mg']),
        float(data['Na'])
    ]

    prediction = predictor.predict_fertility(nutrient_values)
    result_text = predictor.evaluate_fertility(prediction, nutrient_values)
    prediction_percentage = int(round(prediction))
    return render_template('result.html', prediction=prediction_percentage,result_text=result_text, prediction_text=prediction_percentage)

@app.route('/disease')
def disease():
    return render_template('disease.html')

@app.route('/process_image', methods=['POST'])
def process_image():
    if 'imageUpload' not in request.files:
        return "No file uploaded", 400

    file = request.files['imageUpload']
    if file.filename == '':
        return "No selected file", 400

    # Create the uploads directory if it doesn't exist
    upload_dir = os.path.join(app.root_path, 'uploads')
    os.makedirs(upload_dir, exist_ok=True)

    # Save the uploaded file to the uploads directory
    upload_path = os.path.join(upload_dir, file.filename)
    file.save(upload_path)

    # Get the predicted disease
    predicted_disease = predict_disease(upload_path)

    # Render the result template with prediction and image path
    return render_template('Disease_result.html', predicted_disease=predicted_disease, image_path=upload_path)

@app.route('/Comming_soon')
def comming_soon():
    return render_template('comming_soon.html')

try:
    wb = load_workbook('form_data.xlsx')
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    # Add column headers
    ws.append(['Name', 'Email', 'Message'])
    wb.save('form_data.xlsx')
@app.route('/submit_form', methods=['POST'])
def submit_form():
  if request.method == 'POST':
        name = request.form['name']
        email = request.form['mail']
        message = request.form['text']

        # Append data to Excel file
        ws.append([name, email, message])
        wb.save('form_data.xlsx')
        return redirect('/thank_you')
    
@app.route('/thank_you')
def thank_you():
    return 'Thank you for your submission!'

@app.route('/about_us')
def about_us():
    return render_template("about.html")
if __name__ == '__main__':
    app.run(debug=True)
